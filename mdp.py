import gym

import matplotlib.pyplot as plt
import numpy as np
import os
import time
import pandas as pd
from openpyxl import load_workbook
import json

locs = [(0,0), (0,4), (4,0), (4,3)]

def to_json(fn, data):
    with open(fn, 'w') as f:
        json.dump(data, f)
        
def from_json(fn):
    with open(fn) as f:
        return (json.load(f))

def taxi_grid_extract(problem, state):
    grids = []
    state1 = state
    dec = list(problem.decode(state))
    state2 = problem.encode(locs[dec[2]][0], locs[dec[2]][1], 4, dec[3])
    states = [state1, state2]
    for state in states:
        s = [state]
        num = state
        decode = list(problem.decode(state))
        for i in range(decode[0]):
            num -= 100
            s.append(num)            
        num=state
        for m in range(4-decode[0]):
            num += 100
            s.append(num)
                    
        ss = s[:]            
        for i in range(len(ss)):    
            num = ss[i]
            for j in range(decode[1]):
                num -= 20
                s.append(num)        
            num = ss[i]    
            for j in range(4-decode[1]):
                num += 20
                s.append(num)
        list.sort(s)
        grids.append(np.reshape(s, (5,5)))
    return grids

def run_policy(env, policy):
    problem = gym.make(env)
    problem.seed(2019)
    
    rewards = []
    av_reward = []
    ts = []
    outcomes = []
    
    mean_reward = np.nan
    for i_episode in range(1,10001):
        total_rewards = 0
        observation = problem.reset()
        for t in range(5000):
            #problem.render()
            action = policy[observation]
            observation, reward, done, info = problem.step(action)
            total_rewards += reward
            if done:
                if (t+1) >= problem._max_episode_steps:
                     outcomes.append(0)
                     #print("LOSE")
                else:
                     outcomes.append(1)
                     #print("WIN")
                 #print("Episode finished after {} timesteps".format(t+1))
                ts.append(t+1)
                av_reward.append(mean_reward)
                break
        rewards.append(total_rewards)
        if i_episode % 50 == 0:
            mean_reward = np.mean(rewards[i_episode-50:])            
            #print (np.mean(rewards[i_episode-50:]))
            #print('Current avg_reward: %f' % np.mean(av_reward))
    #env.monitor.close()
    #print (np.mean(av_reward))
    return [ts, outcomes, rewards, av_reward]
    

def evaluate_rewards_and_transitions(problem, mutate=False):
    # Enumerate state and action space sizes
    num_states = problem.observation_space.n
    num_actions = problem.action_space.n

    # Intiailize T and R matrices
    R = np.zeros((num_states, num_actions, num_states))
    T = np.zeros((num_states, num_actions, num_states))

    # Iterate over states, actions, and transitions
    for state in range(num_states):
        for action in range(num_actions):
            for transition in problem.env.P[state][action]:
                probability, next_state, reward, done = transition
                R[state, action, next_state] = reward
                T[state, action, next_state] = probability

            # Normalize T across state + action axes
            T[state, action, :] /= np.sum(T[state, action, :])

    # Conditionally mutate and return
    if mutate:
        problem.env.R = R
        problem.env.T = T
    return R, T

#@timing
def value_iteration(problem, R=None, T=None, gamma=0.9, max_iterations=10**6, delta=10**-6):
    """ Runs Value Iteration on a gym problem """
    timesteps = []
    iterations = []
    deltas = []
    gammas = []
    thresholds = []
    results = {}
    
    value_fn = np.zeros(problem.observation_space.n)
    if R is None or T is None:
        R, T = evaluate_rewards_and_transitions(problem)

    for i in range(max_iterations):
        st = time.clock()
        previous_value_fn = value_fn.copy()
        Q = np.einsum('ijk,ijk -> ij', T, R + gamma * value_fn)
        value_fn = np.max(Q, axis=1)
        
        deltas.append(np.max(np.abs(value_fn - previous_value_fn)))
        timesteps.append(time.clock()-st)
        iterations.append(i+1)
        gammas.append(gamma)
        thresholds.append(delta)
        
        if np.max(np.abs(value_fn - previous_value_fn)) < delta:
            break

    # Get and return optimal policy
    policy = np.argmax(Q, axis=1)
    
    results['gamma'] = gammas
    results['threshold'] = thresholds
    results['iteration'] = iterations
    results['timesteps'] = timesteps
    results['delta'] = deltas
    
    return policy, Q, i + 1, results

def encode_policy(policy, shape):
    """ One-hot encodes a policy """
    encoded_policy = np.zeros(shape)
    encoded_policy[np.arange(shape[0]), policy] = 1
    return encoded_policy

#@timing
def policy_iteration(problem, R=None, T=None, gamma=0.9, max_iterations=10**6, delta=10**-6):
    """ Runs Policy Iteration on a gym problem """
    timesteps = []
    iterations = []
    vi_iter = []
    deltas = []
    gammas = []
    thresholds = []
    results = {}
    
    num_spaces = problem.observation_space.n
    num_actions = problem.action_space.n

    # Initialize with a random policy and initial value function
    policy = np.array([problem.action_space.sample() for _ in range(num_spaces)])
    value_fn = np.zeros(num_spaces)

    # Get transitions and rewards
    if R is None or T is None:
        R, T = evaluate_rewards_and_transitions(problem)

    # Iterate and improve policies
    for i in range(max_iterations):
        previous_policy = policy.copy()
        st = time.clock()
        for j in range(max_iterations):
            previous_value_fn = value_fn.copy()
            Q = np.einsum('ijk,ijk -> ij', T, R + gamma * value_fn)
            value_fn = np.sum(encode_policy(policy, (num_spaces, num_actions)) * Q, 1)
            if np.max(np.abs(previous_value_fn - value_fn)) < delta:
                break
            
        vi_iter.append(j+1)
        deltas.append(np.sum(value_fn))
        timesteps.append(time.clock()-st)
        iterations.append(i+1)
        gammas.append(gamma)
        thresholds.append(delta)
        
        Q = np.einsum('ijk,ijk -> ij', T, R + gamma * value_fn)
        policy = np.argmax(Q, axis=1)
        
        if np.array_equal(policy, previous_policy):
            break
        
    results['gamma'] = gammas
    results['threshold'] = thresholds
    results['iteration'] = iterations
    results['vi_iteration'] = vi_iter
    results['timesteps'] = timesteps
    results['delta'] = deltas

    # Return optimal policy
    return policy, Q, i + 1, results

def print_policy(policy, mapping=None, shape=(0,)):
    print (np.array([mapping[action] for action in policy]).reshape(shape))

def run_discrete(environment_name, mapping, shape=None, gamma=0.9, delta=1e-6):
    problem = gym.make(environment_name)
    #problem.seed(2019)
    #print ('== {} =='.format(environment_name))
    #print ('Actions:', problem.env.action_space.n)
    #print ('States:', problem.env.observation_space.n)
    #  print (problem.env.desc)
    #print

    #print ('== Value Iteration ==')
    value_policy, Q_VI, iters, results_VI = value_iteration(problem, gamma=gamma, delta=delta)
    #print ('Iterations:', iters)
    #print
    #print (np.reshape(np.max(Q_VI, axis=1), shape))

    #print ('== Policy Iteration ==')
    policy, Q_PI, iters, results_PI = policy_iteration(problem, gamma=gamma, delta=delta)
    #print ('Iterations:', iters)
    #print
    #print (np.reshape(np.max(Q_PI, axis=1), shape))
    print

    diff = sum([abs(x-y) for x, y in zip(policy.flatten(), value_policy.flatten())])
    #if diff > 0:
        #print ('Discrepancy:', diff)
        #print

    #if shape is not None:
        #print ('== Policy ==')
        #print_policy(policy, mapping, shape)
        #print

    return value_policy, policy, Q_VI, Q_PI, results_VI, results_PI


discounts =[0.1, 0.5, 0.7, 0.8, 0.9, 0.99]
criterias =[1e-1, 1e-2,1e-3,1e-6, 1e-9]


# FROZEN LAKE LARGE
mapping = {0: "L", 1: "D", 2: "R", 3: "U"}
shape = (8, 8)
envname = 'FrozenLake8x8-v0'

#output1 - convergence data
path = './OUTPUT/frozenlake-MDP_Results.xlsx'
dfv = pd.DataFrame()
dfp = pd.DataFrame()
for discount in discounts:
    for criteria in criterias:
        print (envname, "gamma=",discount,"delta=",criteria)
        policy_VI, policy_PI, Q_VI, Q_PI, results_VI, results_PI = run_discrete(envname, mapping, gamma=discount, delta=criteria)
        
        df = pd.DataFrame(results_VI)
        if dfv.empty:
            dfv = df.copy()
        else:
            dfv = dfv.append(df)
            
        df = pd.DataFrame(results_PI)
        if dfp.empty:
            dfp = df.copy()
        else:
            dfp = dfp.append(df)
        #print (dfv)

dfv.to_excel(path, sheet_name='VI')
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book   
dfp.to_excel(writer, sheet_name='PI')
writer.save()
writer.close()       

#output2 - values and policies (full states)
path = './OUTPUT/frozenlake-values_policies_full.json'
data = {}
for discount in discounts:
    res = {}
    for criteria in criterias:
        #print (discount, criteria)
        policy_VI, policy_PI, Q_VI, Q_PI, results_VI, results_PI = run_discrete(envname, mapping, gamma=discount, delta=criteria)
        res[criteria] = {'pol_vi':policy_VI.tolist(), 'pol_pi':policy_PI.tolist(), 'Q_vi':Q_VI.tolist(), 'Q_pi':Q_PI.tolist()}
        #res[criteria] = run_discrete(envname, mapping, gamma=discount, delta=criteria)
        data[discount] = res

to_json(path, data)


#output4 - performance
path = './OUTPUT/frozenlake-performance.xlsx'
dfv = pd.DataFrame()
dfp = pd.DataFrame()
results_vi = {}
results_pi = {}
for discount in discounts:
    res = {}
    for criteria in criterias:  
        print ("preparing output4..")
        res_vi = run_policy(envname, data[discount][criteria]['pol_vi'])
        results_vi['gamma']=discount
        results_vi['delta']=criteria
        results_vi['timestep'] = res_vi[0]
        results_vi['outcomes'] = res_vi[1]
        results_vi['rewards'] = res_vi[2]
        results_vi['av_reward'] = res_vi[3]

        res_pi = run_policy(envname, data[discount][criteria]['pol_pi'])
        results_pi['gamma']=discount
        results_pi['delta']=criteria
        results_pi['timestep'] = res_pi[0]
        results_pi['outcomes'] = res_pi[1]
        results_pi['rewards'] = res_pi[2]
        results_pi['av_reward'] = res_pi[3]
        
        df = pd.DataFrame(results_vi)
        if dfv.empty:
            dfv = df.copy()
        else:
            dfv = dfv.append(df)
            
        df = pd.DataFrame(results_pi)
        if dfp.empty:
            dfp = df.copy()
        else:
            dfp = dfp.append(df)
            
dfv.to_excel(path, sheet_name='VI')
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book   
dfp.to_excel(writer, sheet_name='PI')
writer.save()
writer.close()       


# TAXI
discounts = [0.1, 0.5, 0.7, 0.8, 0.9, 0.99]
criterias = [1e-1, 1e-2,1e-3,1e-6, 1e-9]
mapping = {0: "S", 1: "N", 2: "E", 3: "W", 4: "P", 5: "D"}
envname = 'Taxi-v2'

#output1 - convergence data
path = './OUTPUT/taxi-MDP_Results.xlsx'
dfv = pd.DataFrame()
dfp = pd.DataFrame()
for discount in discounts:
    for criteria in criterias:
        print (envname, "gamma=",discount,"delta=",criteria)
        policy_VI, policy_PI, Q_VI, Q_PI, results_VI, results_PI = run_discrete(envname, mapping, gamma=discount, delta=criteria)
        
        df = pd.DataFrame(results_VI)
        if dfv.empty:
            dfv = df.copy()
        else:
            dfv = dfv.append(df)
            
        df = pd.DataFrame(results_PI)
        if dfp.empty:
            dfp = df.copy()
        else:
            dfp = dfp.append(df)
        #print (dfv)

dfv.to_excel(path, sheet_name='VI')
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book   
dfp.to_excel(writer, sheet_name='PI')
writer.save()
writer.close()       

#output2 - values and policies (full states)
path = './OUTPUT/taxi-values_policies_full.json'
data = {}
for discount in discounts:
    res = {}
    for criteria in criterias:
        #print (discount, criteria)
        policy_VI, policy_PI, Q_VI, Q_PI, results_VI, results_PI = run_discrete(envname, mapping, gamma=discount, delta=criteria)
        res[criteria] = {'pol_vi':policy_VI.tolist(), 'pol_pi':policy_PI.tolist(), 'Q_vi':Q_VI.tolist(), 'Q_pi':Q_PI.tolist()}
        #res[criteria] = run_discrete(envname, mapping, gamma=discount, delta=criteria)
        data[discount] = res

to_json(path, data)


#output3 - values and policies (1 instance)
path = './OUTPUT/taxi-values_policies_instance.json'
problem = gym.make(envname) 
problem.seed(2019)
startstate = problem.reset()
taxi_grid1, taxi_grid2 = taxi_grid_extract(problem, startstate)

tests = {}
for discount in discounts:
    res = {}
    for criteria in criterias:        
        policy_VI_grid1=np.asarray(data[discount][criteria]['pol_vi'])[taxi_grid1.flatten()]
        policy_VI_grid2=np.asarray(data[discount][criteria]['pol_vi'])[taxi_grid2.flatten()]
        policy_PI_grid1=np.asarray(data[discount][criteria]['pol_pi'])[taxi_grid1.flatten()]
        policy_PI_grid2=np.asarray(data[discount][criteria]['pol_pi'])[taxi_grid2.flatten()]
        #print ()
        #print (policy_VI_grid1)
        #print (policy_VI_grid2)
        #print (policy_PI_grid1)
        #print (policy_PI_grid2)
        
        vf_VI_grid1=np.max(data[discount][criteria]['Q_vi'], axis=1)[taxi_grid1.flatten()]
        vf_VI_grid2=np.max(data[discount][criteria]['Q_vi'], axis=1)[taxi_grid2.flatten()]
        vf_PI_grid1=np.max(data[discount][criteria]['Q_pi'], axis=1)[taxi_grid1.flatten()]
        vf_PI_grid2=np.max(data[discount][criteria]['Q_pi'], axis=1)[taxi_grid2.flatten()]
        #print ()
        #print (vf_VI_grid1)
        #print (vf_VI_grid2)
        #print (vf_PI_grid1)
        #print (vf_PI_grid2)
        res[criteria] = {'polvi_grid1':policy_VI_grid1.tolist(), 'polvi_grid2':policy_VI_grid2.tolist(), 
                         'polpi_grid1':policy_PI_grid1.tolist(), 'polpi_grid2':policy_PI_grid2.tolist(),
                         'valvi_grid1':vf_VI_grid1.tolist(), 'valvi_grid2':vf_VI_grid2.tolist(), 
                         'valpi_grid1':vf_PI_grid1.tolist(), 'valpi_grid2':vf_PI_grid2.tolist()}
        
        tests[discount] = res

to_json(path, tests)        


#output4 - performance
path = './OUTPUT/taxi-performance.xlsx'
dfv = pd.DataFrame()
dfp = pd.DataFrame()
results_vi = {}
results_pi = {}
for discount in discounts:
    res = {}
    for criteria in criterias:  
#print ('== Run Value Iteration Policy ==')
        res_vi = run_policy(envname, data[discount][criteria]['pol_vi'])
        results_vi['gamma']=discount
        results_vi['delta']=criteria
        results_vi['timestep'] = res_vi[0]
        results_vi['outcomes'] = res_vi[1]
        results_vi['rewards'] = res_vi[2]
        results_vi['av_reward'] = res_vi[3]

#print ('== Run Policy Iteration Policy ==')
        res_pi = run_policy(envname, data[discount][criteria]['pol_pi'])
        results_pi['gamma']=discount
        results_pi['delta']=criteria
        results_pi['timestep'] = res_pi[0]
        results_pi['outcomes'] = res_pi[1]
        results_pi['rewards'] = res_pi[2]
        results_pi['av_reward'] = res_pi[3]
        
        df = pd.DataFrame(results_vi)
        if dfv.empty:
            dfv = df.copy()
        else:
            dfv = dfv.append(df)
            
        df = pd.DataFrame(results_pi)
        if dfp.empty:
            dfp = df.copy()
        else:
            dfp = dfp.append(df)
            
dfv.to_excel(path, sheet_name='VI')
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book   
dfp.to_excel(writer, sheet_name='PI')
writer.save()
writer.close()       
